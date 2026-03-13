from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import List, Optional, Dict, Any
from datetime import date
from calendar import monthrange
from app.models import get_db
from app.models.schedule import Schedule
from app.models.employee import Employee
from app.models.shift import ShiftType
from app.schemas.schedule import ScheduleCreate, ScheduleUpdate, ScheduleResponse, ScheduleWithDetails, PaginatedResponse
from app.services.scheduler import SchedulerService

router = APIRouter(prefix="/api/schedules", tags=["排班管理"])


@router.get("", response_model=PaginatedResponse[ScheduleWithDetails])
def get_schedules(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    page: int = Query(1, ge=1, description="页码，从1开始"),
    page_size: int = Query(20, ge=1, le=500, description="每页数量，最大500"),
    db: Session = Depends(get_db)
):
    """获取排班列表（支持分页）"""
    query = db.query(Schedule).options(
        joinedload(Schedule.employee),
        joinedload(Schedule.shift_type)
    )

    if start_date:
        query = query.filter(Schedule.date >= start_date)
    if end_date:
        query = query.filter(Schedule.date <= end_date)
    if employee_id:
        query = query.filter(Schedule.employee_id == employee_id)

    # 获取总数
    total = query.count()

    # 分页查询
    skip = (page - 1) * page_size
    schedules = query.order_by(Schedule.date).offset(skip).limit(page_size).all()

    # 添加详细信息
    items = []
    for s in schedules:
        items.append(ScheduleWithDetails(
            id=s.id,
            employee_id=s.employee_id,
            shift_type_id=s.shift_type_id,
            date=s.date,
            employee_name=s.employee.name if s.employee else "",
            shift_name=s.shift_type.name if s.shift_type else "",
            shift_color=s.shift_type.color if s.shift_type else "#409EFF"
        ))

    return PaginatedResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


@router.get("/stats")
def get_schedule_stats(
    type: str = Query("month", description="统计类型: week(周) 或 month(月)"),
    year: int = Query(..., description="年份"),
    week: Optional[int] = Query(None, ge=1, le=53, description="周数，type=week 时必填"),
    month: Optional[int] = Query(None, ge=1, le=12, description="月份，type=month 时必填"),
    db: Session = Depends(get_db)
):
    """获取员工排班统计"""
    try:
        if type == "month":
            if month is None:
                raise HTTPException(status_code=400, detail="月份统计需要提供 month 参数")
            # 计算月份的第一天和最后一天
            first_day = date(year, month, 1)
            _, last_day_num = monthrange(year, month)
            last_day = date(year, month, last_day_num)
        elif type == "week":
            if week is None:
                raise HTTPException(status_code=400, detail="周统计需要提供 week 参数")
            # 计算周的第一天和最后一天（周一为起始）
            first_day = date.fromisocalendar(year, week, 1)
            last_day = date.fromisocalendar(year, week, 7)
        else:
            raise HTTPException(status_code=400, detail="type 参数必须是 week 或 month")

        # 获取所有员工
        employees = db.query(Employee).filter(Employee.is_active == True).all()
        if not employees:
            return {"year": year, "month": month, "employees": []}

        # 获取该月所有排班记录
        schedules = db.query(Schedule).filter(
            Schedule.date >= first_day,
            Schedule.date <= last_day
        ).all()

        # 构建员工排班统计
        employee_stats = []
        for emp in employees:
            # 该员工的排班记录
            emp_schedules = [s for s in schedules if s.employee_id == emp.id]

            # 统计排班天数（按日期去重）
            schedule_dates = set(s.date for s in emp_schedules)

            # 统计班次分布
            shift_distribution = {}
            total_hours = 0
            for s in emp_schedules:
                shift_name = s.shift_type.name if s.shift_type else "未知"
                shift_distribution[shift_name] = shift_distribution.get(shift_name, 0) + 1

                # 计算工作时长
                if s.shift_type and s.shift_type.start_time and s.shift_type.end_time:
                    try:
                        # 解析时间字符串 HH:MM
                        start_parts = s.shift_type.start_time.split(":")
                        end_parts = s.shift_type.end_time.split(":")
                        start_hour = int(start_parts[0])
                        start_min = int(start_parts[1])
                        end_hour = int(end_parts[0])
                        end_min = int(end_parts[1])

                        # 计算时长（分钟）
                        start_total = start_hour * 60 + start_min
                        end_total = end_hour * 60 + end_min
                        duration_minutes = end_total - start_total

                        # 处理跨天班次（结束时间小于开始时间）
                        if duration_minutes < 0:
                            duration_minutes += 24 * 60  # 加上24小时

                        total_hours += duration_minutes / 60
                    except (ValueError, IndexError):
                        pass

            employee_stats.append({
                "employee_id": emp.id,
                "employee_name": emp.name,
                "total_days": len(schedule_dates),
                "total_hours": round(total_hours, 2),
                "shift_distribution": shift_distribution
            })

        # 根据类型返回不同的数据格式
        if type == "month":
            return {
                "type": "month",
                "year": year,
                "month": month,
                "employees": employee_stats
            }
        else:
            return {
                "type": "week",
                "year": year,
                "week": week,
                "employees": employee_stats
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取统计数据失败: {str(e)}")


import csv
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


@router.get("/export")
def export_schedules(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    format: str = "excel",
    db: Session = Depends(get_db)
):
    """导出排班数据"""
    query = db.query(Schedule).options(
        joinedload(Schedule.employee),
        joinedload(Schedule.shift_type)
    )

    if start_date:
        query = query.filter(Schedule.date >= start_date)
    if end_date:
        query = query.filter(Schedule.date <= end_date)

    schedules = query.order_by(Schedule.date, Schedule.shift_type_id).all()

    if format == "csv":
        # 导出 CSV 格式
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["日期", "员工姓名", "员工职位", "班次", "开始时间", "结束时间", "颜色"])

        for s in schedules:
            writer.writerow([
                s.date.strftime("%Y-%m-%d"),
                s.employee.name if s.employee else "",
                s.employee.position if s.employee else "",
                s.shift_type.name if s.shift_type else "",
                s.shift_type.start_time if s.shift_type else "",
                s.shift_type.end_time if s.shift_type else "",
                s.shift_type.color if s.shift_type else ""
            ])

        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=schedules_{start_date}_{end_date}.csv"
            }
        )
    else:
        # 导出 Excel 格式
        wb = Workbook()
        ws = wb.active
        ws.title = "排班表"

        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        headers = ["日期", "员工姓名", "员工职位", "班次", "开始时间", "结束时间", "颜色"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row, s in enumerate(schedules, 2):
            ws.cell(row=row, column=1, value=s.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=s.employee.name if s.employee else "")
            ws.cell(row=row, column=3, value=s.employee.position if s.employee else "")
            ws.cell(row=row, column=4, value=s.shift_type.name if s.shift_type else "")
            ws.cell(row=row, column=5, value=s.shift_type.start_time if s.shift_type else "")
            ws.cell(row=row, column=6, value=s.shift_type.end_time if s.shift_type else "")
            ws.cell(row=row, column=7, value=s.shift_type.color if s.shift_type else "")

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=schedules_{start_date}_{end_date}.xlsx"
            }
        )


@router.get("/{schedule_id}", response_model=ScheduleWithDetails)
def get_schedule(schedule_id: int, db: Session = Depends(get_db)):
    """获取单个排班"""
    schedule = db.query(Schedule).options(
        joinedload(Schedule.employee),
        joinedload(Schedule.shift_type)
    ).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="排班不存在")

    return ScheduleWithDetails(
        id=schedule.id,
        employee_id=schedule.employee_id,
        shift_type_id=schedule.shift_type_id,
        date=schedule.date,
        employee_name=schedule.employee.name if schedule.employee else "",
        shift_name=schedule.shift_type.name if schedule.shift_type else "",
        shift_color=schedule.shift_type.color if schedule.shift_type else "#409EFF"
    )


@router.post("", response_model=ScheduleResponse)
def create_schedule(schedule: ScheduleCreate, db: Session = Depends(get_db)):
    """手动创建排班"""
    db_schedule = Schedule(**schedule.model_dump())
    db.add(db_schedule)
    db.commit()
    db.refresh(db_schedule)
    return db_schedule


@router.put("/{schedule_id}", response_model=ScheduleResponse)
def update_schedule(schedule_id: int, schedule: ScheduleUpdate, db: Session = Depends(get_db)):
    """更新排班"""
    db_schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not db_schedule:
        raise HTTPException(status_code=404, detail="排班不存在")

    for key, value in schedule.model_dump(exclude_unset=True).items():
        setattr(db_schedule, key, value)

    db.commit()
    db.refresh(db_schedule)
    return db_schedule


@router.delete("/{schedule_id}")
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    """删除排班"""
    db_schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not db_schedule:
        raise HTTPException(status_code=404, detail="排班不存在")

    db.delete(db_schedule)
    db.commit()
    return {"message": "排班已删除"}


@router.post("/batch-delete")
def batch_delete_schedules(
    ids: List[int],
    db: Session = Depends(get_db)
):
    """批量删除排班"""
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的排班ID列表")

    # 查询所有要删除的排班
    schedules = db.query(Schedule).filter(Schedule.id.in_(ids)).all()

    if not schedules:
        raise HTTPException(status_code=404, detail="未找到指定的排班记录")

    deleted_count = len(schedules)

    # 批量删除
    for schedule in schedules:
        db.delete(schedule)

    db.commit()

    return {
        "message": f"成功删除 {deleted_count} 条排班记录",
        "deleted_count": deleted_count
    }


class GenerateRequest:
    """生成排班请求"""
    start_date: date
    end_date: date


@router.post("/generate")
def generate_schedule(
    start_date: date = Query(...),
    end_date: date = Query(...),
    team_id: Optional[int] = Query(None, description="班组ID，不指定则使用所有员工"),
    clear_existing: bool = True,
    db: Session = Depends(get_db)
):
    """生成排班"""
    service = SchedulerService(db)

    # 清除现有排班（如果指定了班组，只清除该班组员工在该日期范围的排班）
    if clear_existing:
        if team_id is not None:
            # 获取该班组下的所有员工ID
            employees = db.query(Employee).filter(
                Employee.team_id == team_id
            ).all()
            employee_ids = [emp.id for emp in employees]
            if employee_ids:
                deleted = db.query(Schedule).filter(
                    Schedule.date >= start_date,
                    Schedule.date <= end_date,
                    Schedule.employee_id.in_(employee_ids)
                ).delete()
                db.commit()
            else:
                deleted = 0
        else:
            deleted = service.clear_schedule(start_date, end_date)

    # 生成新排班
    schedules = service.generate_schedule(start_date, end_date, team_id)

    return {
        "message": f"成功生成 {len(schedules)} 条排班记录",
        "schedules": schedules,
        "deleted": deleted if clear_existing else 0,
        "team_id": team_id
    }


@router.delete("/clear")
def clear_schedules(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """清除排班"""
    service = SchedulerService(db)
    if not start_date:
        start_date = date.today()
    if not end_date:
        end_date = date.today()

    deleted = service.clear_schedule(start_date, end_date)
    return {"message": f"已清除 {deleted} 条排班记录"}

