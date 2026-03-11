from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
from app.models import get_db
from app.models.schedule import Schedule
import csv
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/api/schedules", tags=["排班导出"])


@router.get("/export")
def export_schedules(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    format: str = "excel",
    db: Session = Depends(get_db)
):
    """导出排班数据"""
    query = db.query(Schedule)

    if start_date:
        query = query.filter(Schedule.date >= start_date)
    if end_date:
        query = query.filter(Schedule.date <= end_date)

    schedules = query.order_by(Schedule.date, Schedule.shift_type_id).all()

    if format == "csv":
        # 导出 CSV 格式
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["日期", "员工姓名", "员工职位", "班次", "开始时间", "结束时间", "颜色"])

        for s in schedules:
            writer.writerow([
                s.date.strftime("%Y-%m-%d"),
                s.employee.name if s.employee else "",
                s.employee.position if s.employee else "",
                s.shift_type.name if s.shift_type else "",
                s.shift_type.start_time if s.shift_type else "",
                s.shift_type.end_time if s.shift_type else "",
                s.shift_type.color if s.shift_type else ""
            ])

        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=schedules_{start_date}_{end_date}.csv"
            }
        )
    else:
        # 导出 Excel 格式
        wb = Workbook()
        ws = wb.active
        ws.title = "排班表"

        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        headers = ["日期", "员工姓名", "员工职位", "班次", "开始时间", "结束时间", "颜色"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row, s in enumerate(schedules, 2):
            ws.cell(row=row, column=1, value=s.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=s.employee.name if s.employee else "")
            ws.cell(row=row, column=3, value=s.employee.position if s.employee else "")
            ws.cell(row=row, column=4, value=s.shift_type.name if s.shift_type else "")
            ws.cell(row=row, column=5, value=s.shift_type.start_time if s.shift_type else "")
            ws.cell(row=row, column=6, value=s.shift_type.end_time if s.shift_type else "")
            ws.cell(row=row, column=7, value=s.shift_type.color if s.shift_type else "")

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=schedules_{start_date}_{end_date}.xlsx"
            }
        )
